"""
export_excel.py — Generate pharmacist review Excel for FitMate

Merges SymMap raw data + BPOM Indonesian names into a structured Excel file
with dropdowns and validation for pharmacist review.

Output: output/tcm_for_validation.xlsx
Pharmacist saves reviewed file to: validated/tcm_validated.xlsx
"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date

OUTPUT_DIR = Path(__file__).parent / "output"
VALIDATED_DIR = Path(__file__).parent / "validated"
VALIDATED_DIR.mkdir(exist_ok=True)

TOXICITY_LEVELS = '"low,moderate,high,unknown"'
ORGAN_OPTIONS = '"liver,heart,kidney,lung,brain,skin,multiple,none"'
SEVERITY_OPTIONS = '"warning,danger,contraindicated"'


def load_symmap_data() -> pd.DataFrame:
    """Load SymMap scraped herb data."""
    path = OUTPUT_DIR / "symmap_herbs_raw.json"
    if not path.exists():
        print("[warn] SymMap data not found. Run tcm_scraper.py first.")
        return pd.DataFrame()
    return pd.read_json(path)


def load_bpom_data() -> list[dict]:
    """Load scraped BPOM ingredient data."""
    path = OUTPUT_DIR / "bpom_ingredients.json"
    if not path.exists():
        print("[warn] BPOM data not found. Run bpom_scraper.py first.")
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def build_validation_excel(symmap_df: pd.DataFrame, bpom_data: list[dict]):
    """Create pharmacist review Excel with dropdowns and styling."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TCM Validation"

    # Header styling
    header_fill = PatternFill("solid", fgColor="930014")  # Imperial Red
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = [
        ("A", "mandarin_name", "Mandarin Name (锁定)"),
        ("B", "pinyin_name", "Pinyin Name"),
        ("C", "latin_name", "Latin/Scientific Name"),
        ("D", "indonesian_name", "Indonesian Name ← VERIFY"),
        ("E", "is_toxic", "Is Toxic? (TRUE/FALSE) ← DECIDE"),
        ("F", "target_organ", "Target Organ ← SELECT"),
        ("G", "toxicity_level", "Toxicity Level ← SELECT"),
        ("H", "warning_message", "Warning Message (Indonesian) ← WRITE"),
        ("I", "medical_advice", "Medical Advice (Chatbot response) ← WRITE"),
        ("J", "source_reference", "Source (SymMap ID / BPOM ref)"),
        ("K", "validated", "Validated? (TRUE when done) ← CHECK"),
    ]

    for col_letter, _, header_text in headers:
        cell = ws[f"{col_letter}1"]
        cell.value = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[1].height = 40

    # Add data validation dropdowns
    dv_organ = DataValidation(type="list", formula1=ORGAN_OPTIONS, allow_blank=True)
    dv_toxicity = DataValidation(type="list", formula1=TOXICITY_LEVELS, allow_blank=True)
    ws.add_data_validation(dv_organ)
    ws.add_data_validation(dv_toxicity)
    dv_organ.sqref = "F2:F1000"
    dv_toxicity.sqref = "G2:G1000"

    # Populate with SymMap data (first 200 rows for prototype)
    row = 2
    if not symmap_df.empty:
        for _, herb in symmap_df.head(200).iterrows():
            ws.cell(row=row, column=1, value=str(herb.get("Chinese_name", herb.get("chinese_name", ""))))
            ws.cell(row=row, column=2, value=str(herb.get("Pinyin_name", herb.get("pinyin_name", ""))))
            ws.cell(row=row, column=3, value=str(herb.get("Latin_name", herb.get("latin_name", ""))))
            ws.cell(row=row, column=10, value=str(herb.get("SMHB_id", herb.get("symmap_id", f"SMHB-{row}"))))
            ws.cell(row=row, column=11, value=False)  # Not yet validated
            row += 1

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-column widths
    for col_idx in range(1, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    output_path = OUTPUT_DIR / "tcm_for_validation.xlsx"
    wb.save(output_path)
    print(f"[ok] Generated pharmacist review Excel: {output_path}")
    print(f"[next] Share {output_path} with pharmacy team for validation")
    print(f"[next] Pharmacy team saves validated file to: {VALIDATED_DIR}/tcm_validated.xlsx")


def main():
    print("=== FitMate Excel Export: Pharmacist Review Generator ===\n")
    symmap_df = load_symmap_data()
    bpom_data = load_bpom_data()
    build_validation_excel(symmap_df, bpom_data)
    print(f"\nGenerated: {date.today()}")


if __name__ == "__main__":
    main()
